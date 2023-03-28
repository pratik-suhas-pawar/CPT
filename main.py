import threading
import multiprocessing
import time
from datetime import datetime
import os
from git_data import GitData
from tkinter import Tk, LabelFrame, Label, Button, Entry, filedialog, StringVar
from tkinter.ttk import Progressbar, Style
from PIL import Image, ImageTk
import openpyxl

with open(f"data.dat", "r") as io:
    raw = io.read().split(',')

width, height = 1366, 768
g_data = GitData()
date, month_year = str(int(datetime.now().strftime("%d"))), datetime.now().strftime("%B %Y").split(" ")


class PTG:
    def __init__(self, window):

        self.data = {}
        self.data_name = []
        self.main = window
        self.frame_1 = LabelFrame(self.main, bg="white", fg="black", bd=0)
        self.main.geometry(newGeometry=f"{width}x{height}+{int((self.main.winfo_screenwidth() - width) / 2)}+"
                                       f"{int((self.main.winfo_screenheight() - height) / 2) - 20}")
        self.main.title("C P T")

        self.main.bind('<Button-1>', self.click)
        self.notification = Label(self.main, fg="black", bg="white", font=("Segoe UI", 12), relief="solid",
                                  bd=0, anchor="s")
        self.notification.place(x=width - 100, y=height - 20, width=100, height=20)
        self.sync()
        self.main.protocol('WM_DELETE_WINDOW', self.exit)

    def sync(self):
        if f"{month_year[0] + ' ' + date + ', ' + month_year[1]}.xml" in os.listdir("essential/"):
            print("Syncing from database")
            self.notification.config(text="Local")
            with open(f"essential/{month_year[0] + ' ' + date + ', ' + month_year[1]}.xml", "r") as file:
                data = file.read()

            for name_comit in data.split(",")[:-1]:
                name, comit = name_comit.split(':')[0], int(name_comit.split(':')[-1])
                self.data[name] = comit
                self.data_name.append(name)

        else:
            t1 = threading.Thread(target=self.fetch_data)
            t1.start()

    def fetch_data(self):
        self.notification.config(text="Syncing...")
        dataframe = openpyxl.load_workbook(raw[0])
        dataframe1 = dataframe.active

        data_base = open(f"essential/{month_year[0] + ' ' + date + ', ' + month_year[1]}.xml", "a")
        for i in dataframe1.iter_rows(2, dataframe1.max_row):
            name, usr_id = i[0].value, i[1].value
            comit = g_data.get_today(usr_id)
            self.data[name] = comit
            self.data_name.append(name)
            data_base.write(f"{name}:{comit},")
        print(self.data)
        data_base.close()
        self.notification.config(text="Updated")
        self.notification.update()

        def progress(self):
        self.frame_1 = LabelFrame(self.main, bg="white", fg="black", bd=0)
        # self.cnt = 1
        self.page = 0

        def close_progress():
            self.frame_1.destroy()

        self.frame_1.place(x=0, y=0, width=width, height=height)
        Button(self.frame_1, text="↼", fg="black", bg="white", font=("calibre", 25), foreground="black",
               background="white", relief="solid", bd=0, command=close_progress).place(x=0, y=0, width=50, height=50)
        style = Style()
        style.theme_use('alt')
        style.configure("pg_1.Horizontal.TProgressbar", troughcolor="white",
                        bordercolor="black", background="orange", lightcolor="white",
                        darkcolor="white")
        style.configure("pg_2.Horizontal.TProgressbar", troughcolor="white",
                        bordercolor="black", background="orange", lightcolor="white",
                        darkcolor="white")
        style.configure("pg_3.Horizontal.TProgressbar", troughcolor="white",
                        bordercolor="black", background="orange", lightcolor="white",
                        darkcolor="white")
        style.configure("pg_4.Horizontal.TProgressbar", troughcolor="white",
                        bordercolor="black", background="orange", lightcolor="white",
                        darkcolor="white")
        style.configure("pg_5.Horizontal.TProgressbar", troughcolor="white",
                        bordercolor="black", background="orange", lightcolor="white",
                        darkcolor="white")
        style.configure("pg_6.Horizontal.TProgressbar", troughcolor="white",
                        bordercolor="black", background="orange", lightcolor="white",
                        darkcolor="white")
        style.configure("pg_7.Horizontal.TProgressbar", troughcolor="white",
                        bordercolor="black", background="orange", lightcolor="white",
                        darkcolor="white")
        style.configure("pg_8.Horizontal.TProgressbar", troughcolor="white",
                        bordercolor="black", background="orange", lightcolor="white",
                        darkcolor="white")
        style.configure("pg_9.Horizontal.TProgressbar", troughcolor="white",
                        bordercolor="black", background="orange", lightcolor="white",
                        darkcolor="white")
        style.configure("pg_10.Horizontal.TProgressbar", troughcolor="white",
                        bordercolor="black", background="orange", lightcolor="white",
                        darkcolor="white")
        style.configure("pg_11.Horizontal.TProgressbar", troughcolor="white",
                        bordercolor="black", background="orange", lightcolor="white",
                        darkcolor="white")
        style.configure("pg_12.Horizontal.TProgressbar", troughcolor="white",
                        bordercolor="black", background="orange", lightcolor="white",
                        darkcolor="white")
        style.configure("pg_13.Horizontal.TProgressbar", troughcolor="white",
                        bordercolor="black", background="orange", lightcolor="white",
                        darkcolor="white")
        style.configure("pg_14.Horizontal.TProgressbar", troughcolor="white",
                        bordercolor="black", background="orange", lightcolor="white",
                        darkcolor="white")

        style.configure("pgr_1.Horizontal.TProgressbar", troughcolor="white",
                        bordercolor="black", background="orange", lightcolor="white",
                        darkcolor="white")
        style.configure("pgr_2.Horizontal.TProgressbar", troughcolor="white",
                        bordercolor="black", background="orange", lightcolor="white",
                        darkcolor="white")
        style.configure("pgr_3.Horizontal.TProgressbar", troughcolor="white",
                        bordercolor="black", background="orange", lightcolor="white",
                        darkcolor="white")
        style.configure("pgr_4.Horizontal.TProgressbar", troughcolor="white",
                        bordercolor="black", background="orange", lightcolor="white",
                        darkcolor="white")
        style.configure("pgr_5.Horizontal.TProgressbar", troughcolor="white",
                        bordercolor="black", background="orange", lightcolor="white",
                        darkcolor="white")
        style.configure("pgr_6.Horizontal.TProgressbar", troughcolor="white",
                        bordercolor="black", background="orange", lightcolor="white",
                        darkcolor="white")
        style.configure("pgr_7.Horizontal.TProgressbar", troughcolor="white",
                        bordercolor="black", background="orange", lightcolor="white",
                        darkcolor="white")
        style.configure("pgr_8.Horizontal.TProgressbar", troughcolor="white",
                        bordercolor="black", background="orange", lightcolor="white",
                        darkcolor="white")
        style.configure("pgr_9.Horizontal.TProgressbar", troughcolor="white",
                        bordercolor="black", background="orange", lightcolor="white",
                        darkcolor="white")
        style.configure("pgr_10.Horizontal.TProgressbar", troughcolor="white",
                        bordercolor="black", background="orange", lightcolor="white",
                        darkcolor="white")
        style.configure("pgr_11.Horizontal.TProgressbar", troughcolor="white",
                        bordercolor="black", background="orange", lightcolor="white",
                        darkcolor="white")
        style.configure("pgr_12.Horizontal.TProgressbar", troughcolor="white",
                        bordercolor="black", background="orange", lightcolor="white",
                        darkcolor="white")
        style.configure("pgr_13.Horizontal.TProgressbar", troughcolor="white",
                        bordercolor="black", background="orange", lightcolor="white",
                        darkcolor="white")
        style.configure("pgr_14.Horizontal.TProgressbar", troughcolor="white",
                        bordercolor="black", background="orange", lightcolor="white",
                        darkcolor="white")

        pg_1 = Progressbar(self.frame_1, style="pg_1.Horizontal.TProgressbar")
        pg_1.place(x=308, y=50, width=300, height=30)
        pg_2 = Progressbar(self.frame_1, style="pg_2.Horizontal.TProgressbar")
        pg_2.place(x=308, y=100, width=300, height=30)
        pg_3 = Progressbar(self.frame_1, style="pg_3.Horizontal.TProgressbar")
        pg_3.place(x=308, y=150, width=300, height=30)
        pg_4 = Progressbar(self.frame_1, style="pg_4.Horizontal.TProgressbar")
        pg_4.place(x=308, y=200, width=300, height=30)
        pg_5 = Progressbar(self.frame_1, style="pg_5.Horizontal.TProgressbar")
        pg_5.place(x=308, y=250, width=300, height=30)
        pg_6 = Progressbar(self.frame_1, style="pg_6.Horizontal.TProgressbar")
        pg_6.place(x=308, y=300, width=300, height=30)
        pg_7 = Progressbar(self.frame_1, style="pg_7.Horizontal.TProgressbar")
        pg_7.place(x=308, y=350, width=300, height=30)
        pg_8 = Progressbar(self.frame_1, style="pg_8.Horizontal.TProgressbar")
        pg_8.place(x=308, y=400, width=300, height=30)
        pg_9 = Progressbar(self.frame_1, style="pg_9.Horizontal.TProgressbar")
        pg_9.place(x=308, y=450, width=300, height=30)
        pg_10 = Progressbar(self.frame_1, style="pg_10.Horizontal.TProgressbar")
        pg_10.place(x=308, y=500, width=300, height=30)
        pg_11 = Progressbar(self.frame_1, style="pg_11.Horizontal.TProgressbar")
        pg_11.place(x=308, y=550, width=300, height=30)
        pg_12 = Progressbar(self.frame_1, style="pg_12.Horizontal.TProgressbar")
        pg_12.place(x=308, y=600, width=300, height=30)
        pg_13 = Progressbar(self.frame_1, style="pg_13.Horizontal.TProgressbar")
        pg_13.place(x=308, y=650, width=300, height=30)
        pg_14 = Progressbar(self.frame_1, style="pg_14.Horizontal.TProgressbar")
        pg_14.place(x=308, y=700, width=300, height=30)

        pgr_1 = Progressbar(self.frame_1, style="pgr_1.Horizontal.TProgressbar")
        pgr_1.place(x=956, y=50, width=300, height=30)
        pgr_2 = Progressbar(self.frame_1, style="pgr_2.Horizontal.TProgressbar")
        pgr_2.place(x=956, y=100, width=300, height=30)
        pgr_3 = Progressbar(self.frame_1, style="pgr_3.Horizontal.TProgressbar")
        pgr_3.place(x=956, y=150, width=300, height=30)
        pgr_4 = Progressbar(self.frame_1, style="pgr_4.Horizontal.TProgressbar")
        pgr_4.place(x=956, y=200, width=300, height=30)
        pgr_5 = Progressbar(self.frame_1, style="pgr_5.Horizontal.TProgressbar")
        pgr_5.place(x=956, y=250, width=300, height=30)
        pgr_6 = Progressbar(self.frame_1, style="pgr_6.Horizontal.TProgressbar")
        pgr_6.place(x=956, y=300, width=300, height=30)
        pgr_7 = Progressbar(self.frame_1, style="pgr_7.Horizontal.TProgressbar")
        pgr_7.place(x=956, y=350, width=300, height=30)
        pgr_8 = Progressbar(self.frame_1, style="pgr_8.Horizontal.TProgressbar")
        pgr_8.place(x=956, y=400, width=300, height=30)
        pgr_9 = Progressbar(self.frame_1, style="pgr_9.Horizontal.TProgressbar")
        pgr_9.place(x=956, y=450, width=300, height=30)
        pgr_10 = Progressbar(self.frame_1, style="pgr_10.Horizontal.TProgressbar")
        pgr_10.place(x=956, y=500, width=300, height=30)
        pgr_11 = Progressbar(self.frame_1, style="pgr_11.Horizontal.TProgressbar")
        pgr_11.place(x=956, y=550, width=300, height=30)
        pgr_12 = Progressbar(self.frame_1, style="pgr_12.Horizontal.TProgressbar")
        pgr_12.place(x=956, y=600, width=300, height=30)
        pgr_13 = Progressbar(self.frame_1, style="pgr_13.Horizontal.TProgressbar")
        pgr_13.place(x=956, y=650, width=300, height=30)
        pgr_14 = Progressbar(self.frame_1, style="pgr_14.Horizontal.TProgressbar")
        pgr_14.place(x=956, y=700, width=300, height=30)

        lbl_1 = Label(self.frame_1, bg="white", fg="black", font=("Segoe UI", 20))
        lbl_1.place(x=8, y=50, width=300, height=30)
        lbl_2 = Label(self.frame_1, bg="white", fg="black", font=("Segoe UI", 20))
        lbl_2.place(x=8, y=100, width=300, height=30)
        lbl_3 = Label(self.frame_1, bg="white", fg="black", font=("Segoe UI", 20))
        lbl_3.place(x=8, y=150, width=300, height=30)
        lbl_4 = Label(self.frame_1, bg="white", fg="black", font=("Segoe UI", 20))
        lbl_4.place(x=8, y=200, width=300, height=30)
        lbl_5 = Label(self.frame_1, bg="white", fg="black", font=("Segoe UI", 20))
        lbl_5.place(x=8, y=250, width=300, height=30)
        lbl_6 = Label(self.frame_1, bg="white", fg="black", font=("Segoe UI", 20))
        lbl_6.place(x=8, y=300, width=300, height=30)
        lbl_7 = Label(self.frame_1, bg="white", fg="black", font=("Segoe UI", 20))
        lbl_7.place(x=8, y=350, width=300, height=30)
        lbl_8 = Label(self.frame_1, bg="white", fg="black", font=("Segoe UI", 20))
        lbl_8.place(x=8, y=400, width=300, height=30)
        lbl_9 = Label(self.frame_1, bg="white", fg="black", font=("Segoe UI", 20))
        lbl_9.place(x=8, y=450, width=300, height=30)
        lbl_10 = Label(self.frame_1, bg="white", fg="black", font=("Segoe UI", 20))
        lbl_10.place(x=8, y=500, width=300, height=30)
        lbl_11 = Label(self.frame_1, bg="white", fg="black", font=("Segoe UI", 20))
        lbl_11.place(x=8, y=550, width=300, height=30)
        lbl_12 = Label(self.frame_1, bg="white", fg="black", font=("Segoe UI", 20))
        lbl_12.place(x=8, y=600, width=300, height=30)
        lbl_13 = Label(self.frame_1, bg="white", fg="black", font=("Segoe UI", 20))
        lbl_13.place(x=8, y=650, width=300, height=30)
        lbl_14 = Label(self.frame_1, bg="white", fg="black", font=("Segoe UI", 20))
        lbl_14.place(x=8, y=700, width=300, height=30)

        lbr_1 = Label(self.frame_1, bg="white", fg="black", font=("Segoe UI", 20))
        lbr_1.place(x=624, y=50, width=300, height=30)
        lbr_2 = Label(self.frame_1, bg="white", fg="black", font=("Segoe UI", 20))
        lbr_2.place(x=624, y=100, width=300, height=30)
        lbr_3 = Label(self.frame_1, bg="white", fg="black", font=("Segoe UI", 20))
        lbr_3.place(x=624, y=150, width=300, height=30)
        lbr_4 = Label(self.frame_1, bg="white", fg="black", font=("Segoe UI", 20))
        lbr_4.place(x=624, y=200, width=300, height=30)
        lbr_5 = Label(self.frame_1, bg="white", fg="black", font=("Segoe UI", 20))
        lbr_5.place(x=624, y=250, width=300, height=30)
        lbr_6 = Label(self.frame_1, bg="white", fg="black", font=("Segoe UI", 20))
        lbr_6.place(x=624, y=300, width=300, height=30)
        lbr_7 = Label(self.frame_1, bg="white", fg="black", font=("Segoe UI", 20))
        lbr_7.place(x=624, y=350, width=300, height=30)
        lbr_8 = Label(self.frame_1, bg="white", fg="black", font=("Segoe UI", 20))
        lbr_8.place(x=624, y=400, width=300, height=30)
        lbr_9 = Label(self.frame_1, bg="white", fg="black", font=("Segoe UI", 20))
        lbr_9.place(x=624, y=450, width=300, height=30)
        lbr_10 = Label(self.frame_1, bg="white", fg="black", font=("Segoe UI", 20))
        lbr_10.place(x=624, y=500, width=300, height=30)
        lbr_11 = Label(self.frame_1, bg="white", fg="black", font=("Segoe UI", 20))
        lbr_11.place(x=624, y=550, width=300, height=30)
        lbr_12 = Label(self.frame_1, bg="white", fg="black", font=("Segoe UI", 20))
        lbr_12.place(x=624, y=600, width=300, height=30)
        lbr_13 = Label(self.frame_1, bg="white", fg="black", font=("Segoe UI", 20))
        lbr_13.place(x=624, y=650, width=300, height=30)
        lbr_14 = Label(self.frame_1, bg="white", fg="black", font=("Segoe UI", 20))
        lbr_14.place(x=624, y=700, width=300, height=30)

        def show_data(val: int = 0, std_name=""):
            # std_name = std_name + " (" + str(val) + ") "
            if 0 < val <= 4:
                color = "#02231c"
            elif 4 < val <= 8:
                color = "#004d25"
            elif 8 < val <= 15:
                color = "#11823b"
            elif 15 < val <= 20:
                color = "#48bf53"
            else:
                color = "#91f086"
            if 0 <= self.cnt <= 14:
                pg_name = "pg_" + str(self.cnt)
            else:
                pg_name = "pgr_" + str(self.cnt - 14)
            style.configure(f"{pg_name}.Horizontal.TProgressbar", background=color)

            if pg_name == "pg_1":
                pg_1['value'] = val
                lbl_1.config(text=std_name)
            elif pg_name == "pg_2":
                pg_2["value"] = val
                lbl_2.config(text=std_name)
            elif pg_name == "pg_3":
                pg_3["value"] = val
                lbl_3.config(text=std_name)
            elif pg_name == "pg_4":
                pg_4["value"] = val
                lbl_4.config(text=std_name)
            elif pg_name == "pg_5":
                pg_5["value"] = val
                lbl_5.config(text=std_name)
            elif pg_name == "pg_6":
                pg_6["value"] = val
                lbl_6.config(text=std_name)
            elif pg_name == "pg_7":
                pg_7["value"] = val
                lbl_7.config(text=std_name)
            elif pg_name == "pg_8":
                pg_8["value"] = val
                lbl_8.config(text=std_name)
            elif pg_name == "pg_9":
                pg_9["value"] = val
                lbl_9.config(text=std_name)
            elif pg_name == "pg_10":
                pg_10["value"] = val
                lbl_10.config(text=std_name)
            elif pg_name == "pg_11":
                pg_11["value"] = val
                lbl_11.config(text=std_name)
            elif pg_name == "pg_12":
                pg_12["value"] = val
                lbl_12.config(text=std_name)
            elif pg_name == "pg_13":
                pg_13["value"] = val
                lbl_13.config(text=std_name)
            elif pg_name == "pg_14":
                pg_14["value"] = val
                lbl_14.config(text=std_name)
            elif pg_name == "pgr_1":
                pgr_1["value"] = val
                lbr_1.config(text=std_name)
            elif pg_name == "pgr_2":
                pgr_2["value"] = val
                lbr_2.config(text=std_name)
            elif pg_name == "pgr_3":
                pgr_3["value"] = val
                lbr_3.config(text=std_name)
            elif pg_name == "pgr_4":
                pgr_4["value"] = val
                lbr_4.config(text=std_name)
            elif pg_name == "pgr_5":
                pgr_5["value"] = val
                lbr_5.config(text=std_name)
            elif pg_name == "pgr_6":
                pgr_6["value"] = val
                lbr_6.config(text=std_name)
            elif pg_name == "pgr_7":
                pgr_7["value"] = val
                lbr_7.config(text=std_name)
            elif pg_name == "pgr_8":
                pgr_8["value"] = val
                lbr_8.config(text=std_name)
            elif pg_name == "pgr_9":
                pgr_9["value"] = val
                lbr_9.config(text=std_name)
            elif pg_name == "pgr_10":
                pgr_10["value"] = val
                lbr_10.config(text=std_name)
            elif pg_name == "pgr_11":
                pgr_11["value"] = val
                lbr_11.config(text=std_name)
            elif pg_name == "pgr_12":
                pgr_12["value"] = val
                lbr_12.config(text=std_name)
            elif pg_name == "pgr_13":
                pgr_13["value"] = val
                lbr_13.config(text=std_name)
            elif pg_name == "pgr_14":
                pgr_14["value"] = val
                lbr_14.config(text=std_name)

            self.cnt = self.cnt + 1

        def show():
            self.cnt = 1
            for name in self.data_name[0 + (28 * self.page): 28 + (28 * self.page)]:
                contrib = self.data[name]
                show_data(contrib, name)

            for _ in range(28 - len(self.data_name[0 + (28 * self.page): 28 + (28 * self.page)])):
                show_data(0, "")

        show()

        def page_back():
            if self.page > 0:
                self.page = self.page - 1
            self.pg_num_lbl.config(text=self.page)
            show()

        def page_front():
            if self.page < int(len(self.data_name) / 28):
                self.page = self.page + 1
            self.pg_num_lbl.config(text=self.page)

            show()

        Button(self.frame_1, text=" < ", font=("courier new", 15), fg="black", bg="white", relief="solid", bd=0,
               command=page_back).place(x=int((width / 2) - 50), width=20, y=height - 30)

        self.pg_num_lbl = Label(self.frame_1, font=("Segoe UI", 15), fg="black", bg="white", text=self.page,
                                relief="solid", bd=1)
        self.pg_num_lbl.place(x=int((width / 2) - 20), width=40, y=height - 30)

        Button(self.frame_1, text=" > ", font=("courier new", 15), fg="black", bg="white", relief="solid", bd=0,
               command=page_front).place(x=int((width / 2) + 30), width=20, y=height - 30)

    def settings(self):
        se_file, te_file, be_file = StringVar(), StringVar(), StringVar()
        se_file.set(raw[0])
        te_file.set(raw[1])
        be_file.set(raw[2])

        self.frame_2 = LabelFrame(self.main, bg="white", fg="black", bd=0)

        def close_settings(a=None):
            self.frame_2.destroy()

        self.frame_2.place(x=0, y=0, width=width, height=height)
        Button(self.frame_2, text="↼", fg="black", bg="white", font=("calibre", 25), foreground="black",
               background="white", relief="solid", bd=0, command=close_settings).place(x=0, y=0, width=50, height=50)

        box = LabelFrame(self.frame_2, text=" Settings ", foreground="black", background="white", font=("Segoe UI", 20),
                         relief="ridge", bd=2)
        box.place(x=300, y=50, width=width - 600, height=height - 100)

        Label(box, text="SE Git ID: ", font=("Segoe UI", 20), relief="solid", bd=0, fg="black",
              bg="white").place(x=100, y=30, )
        Label(box, text="TE Git ID: ", font=("Segoe UI", 20), relief="solid", bd=0, fg="black",
              bg="white").place(x=100, y=80, )
        Label(box, text="BE Git ID: ", font=("Segoe UI", 20), relief="solid", bd=0, fg="black",
              bg="white").place(x=100, y=130, )

        Label(box, text="Last Sync: " + month_year[0] + ' ' + date + ', ' + month_year[1], font=("Segoe UI", 20), relief="solid", bd=0, fg="black",
              bg="white").place(x=100, y=height - 100 - 70, )

        Entry(box, textvariable=se_file, relief="solid", bd=1, fg="black", bg="white", font=("Segoe UI", 15),
              state="normal").place(x=200, y=30, width=400)
        Entry(box, textvariable=te_file, relief="solid", bd=1, fg="black", bg="white", font=("Segoe UI", 15),
              state="normal").place(x=200, y=80, width=400)
        Entry(box, textvariable=be_file, relief="solid", bd=1, fg="black", bg="white", font=("Segoe UI", 15),
              state="normal").place(x=200, y=130, width=400)

        def get_sec():
            file = filedialog.askopenfilename(filetypes=[('Excel', '*.xlsx')], defaultextension=".xlsx",
                                              title="SE Git Excel", )
            if file is not None:
                se_file.set(file)
                raw[0] = file

        def get_tec():
            file = filedialog.askopenfilename(filetypes=[('Excel', '*.xlsx')], defaultextension=".xlsx",
                                              title="TE Git Excel", )
            if file is not None:
                te_file.set(file)
                raw[1] = file

        def get_bec():
            file = filedialog.askopenfilename(filetypes=[('Excel', '*.xlsx')], defaultextension=".xlsx",
                                              title="BE Git Excel", )
            if file is not None:
                be_file.set(file)
                raw[2] = file

        Button(box, text="Browse", font=("Segoe UI", 15), relief="solid", bd=1, bg="white", command=get_sec,
               fg="black").place(x=620, y=30, height=30)
        Button(box, text="Browse", font=("Segoe UI", 15), relief="solid", bd=1, bg="white", command=get_tec,
               fg="black").place(x=620, y=80, height=30)
        Button(box, text="Browse", font=("Segoe UI", 15), relief="solid", bd=1, bg="white", command=get_bec,
               fg="black").place(x=620, y=130, height=30)

    def click(self, event):
        if 557 > event.x > 457 and 270 < event.y < 370:
            self.progress()
        elif 557 > event.x > 457 and 435 < event.y < 535:
            self.sync()
        elif 734 > event.x > 634 and 270 < event.y < 370:
            print("event 3 triggered")
        elif 734 > event.x > 634 and 435 < event.y < 535:
            print("event 4 triggered")
        elif 910 > event.x > 810 and 270 < event.y < 370:
            print("event 5 triggered")
        elif 910 > event.x > 810 and 435 < event.y < 535:
            self.settings()
        return

    def exit(self):
        with open("data.dat", "w") as save:
            for i in raw:
                save.write(i + ",")
        self.main.destroy()


if __name__ == "__main__":
    win = Tk()
    win.resizable(False, False)
    bg_img = ImageTk.PhotoImage(Image.open("bg.dat"))
    bg = Label(win, image=bg_img)
    bg.pack()

    PTG(win)
    win.mainloop()
